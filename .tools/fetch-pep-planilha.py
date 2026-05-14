#!/usr/bin/env python
"""
Importa dados da planilha CONTROLE DE ANÁLISES (Google Drive — .xlsx)
e mescla com pep-history.json, adicionando casos não capturados pelo
build-pep-history.py (rotas COMPLIANCE_BACEN, REPROCESSAMENTO, etc.).

Estrutura da aba PEP ONB:
  [3]  ID CADASTRO         → draft_membership_id
  [1]  INÍCIO DA ANÁLISE   → pld_entered_at
  [13] CONCLUSÃO DA ANÁLISE → decision_at
  [14] RECOMENDAÇÃO DE COMPLIANCE/PLD → "SEM OBJEÇÃO" | "COM OBJEÇÃO" | "REPROCESSAMENTO"
  [15] PARECER ANALISTA    → "APROVADO" | "REPROVADO" | (vazio = em_andamento)
  [23] COMPETÊNCIA         → mês por extenso ("ABRIL", "MARÇO", ...)
  [24] ANO                 → ano numérico (2026.0)

Mapeamento de status:
  PARECER "APROVADO"         → aprovado
  PARECER "REPROVADO"        → reprovado
  RECOMENDAÇÃO "REPROCESSAMENTO"/"REPROCESSAR" → falso_positivo
  (vazio / outros)           → em_andamento

Configuração .env:
  GOOGLE_SERVICE_ACCOUNT_JSON=/path/to/service-account.json
  PLANILHA_ID=1-esncmUBoGGpY-Q903OI0u8Xx5PD8EzA

Uso:
  python fetch-pep-planilha.py           # mescla com pep-history
  python fetch-pep-planilha.py --dry-run # mostra o que seria importado
  python fetch-pep-planilha.py --mes 4 --ano 2026  # filtrar por mês/ano
"""

import io
import json
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parents[2] / ".env")

ROOT = Path(__file__).resolve().parents[1]
PEP_HISTORY_PATH = ROOT / "src" / "data" / "pep-history.json"

PLANILHA_FILE_ID = os.environ.get("PLANILHA_ID", "1-esncmUBoGGpY-Q903OI0u8Xx5PD8EzA")
ABA_PEP_ONB = "PEP ONB"

MESES_PT = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "ABRIL": 4,
    "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8,
    "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}


def _baixar_xlsx() -> bytes:
    """Baixa o arquivo .xlsx do Google Drive via service account."""
    import requests
    import google.oauth2.service_account as sa_mod
    from google.auth.transport.requests import Request as GRequest

    sa_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not sa_path or not Path(sa_path).exists():
        raise RuntimeError(
            f"Arquivo de service account não encontrado: {sa_path!r}. "
            "Configure GOOGLE_SERVICE_ACCOUNT_JSON no .env"
        )

    creds = sa_mod.Credentials.from_service_account_file(
        sa_path,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    creds.refresh(GRequest())

    resp = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{PLANILHA_FILE_ID}?alt=media",
        headers={"Authorization": f"Bearer {creds.token}"},
        timeout=60,
    )
    resp.raise_for_status()
    print(f"  Download OK: {len(resp.content):,} bytes")
    return resp.content


def _ler_aba(xlsx_bytes: bytes) -> tuple[list[str], list[tuple]]:
    """Carrega a aba PEP ONB e retorna (headers, rows)."""
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    if ABA_PEP_ONB not in wb.sheetnames:
        raise RuntimeError(
            f"Aba '{ABA_PEP_ONB}' não encontrada. Abas: {wb.sheetnames}"
        )

    ws = wb[ABA_PEP_ONB]
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def _map_status(parecer: str | None, recomendacao: str | None) -> str:
    """Converte os campos da planilha em status_pepito."""
    p = (parecer or "").strip().upper()
    r = (recomendacao or "").strip().upper()

    if p == "APROVADO":
        return "aprovado"
    if p == "REPROVADO":
        return "reprovado"
    if r in ("REPROCESSAMENTO", "REPROCESSAR"):
        return "falso_positivo"
    # SEM OBJEÇÃO sem parecer final = aprovado aguardando formalização
    if r == "SEM OBJEÇÃO":
        return "aprovado"
    if r == "COM OBJEÇÃO":
        return "reprovado"
    return "em_andamento"


def processar(rows: list[tuple], filtro_mes: int | None = None, filtro_ano: int | None = None) -> list[dict]:
    """
    Converte as linhas da aba PEP ONB em items pep-history.
    Aplica filtro de mês/ano se fornecido.
    """
    items = []
    for row in rows:
        if len(row) < 25:
            continue

        # Filtrar por competência (mês/ano)
        comp_raw = str(row[23]).strip().upper() if row[23] else ""
        ano_raw = row[24]
        try:
            ano = int(float(ano_raw))
        except (TypeError, ValueError):
            continue  # pula linhas sem ano

        mes_num = MESES_PT.get(comp_raw, 0)
        if not mes_num:
            continue  # pula linhas sem mês legível

        if filtro_mes and mes_num != filtro_mes:
            continue
        if filtro_ano and ano != filtro_ano:
            continue

        draft_id = str(row[3]).strip() if row[3] else None
        if not draft_id or draft_id == "None":
            continue

        parecer = row[15]
        recomendacao = row[14]
        status_pepito = _map_status(parecer, recomendacao)

        # Datas
        conclusao = row[13]
        decision_at = conclusao.isoformat() if isinstance(conclusao, datetime) else str(conclusao) if conclusao else None
        # pld_entered_at usa o mês de COMPETÊNCIA (quando o caso foi analisado pela equipe),
        # não o INÍCIO DA ANÁLISE (que pode ser data de abertura do cadastro, mês anterior).
        pld_entered_at = f"{ano}-{mes_num:02d}-01T00:00:00"

        # Tipo PEP
        cat_pep = str(row[11]).upper() if row[11] else ""
        tipo_pep = "titular" if "TITULAR" in cat_pep else ("relacionado" if cat_pep else None)

        items.append({
            "draft_membership_id": draft_id,
            "status_pepito": status_pepito,
            "raw_status": f"PLANILHA_{str(parecer or recomendacao or 'VAZIO').upper()[:20]}",
            "motivo": str(recomendacao) if status_pepito == "reprovado" and recomendacao else None,
            "motivo_label": str(recomendacao) if recomendacao else "",
            "tipo_pep": tipo_pep,
            "ds_vinculo": None,
            "decision_at": decision_at,
            "pld_entered_at": pld_entered_at,
            # competencia_at = mês de competência da planilha (quando o time analisou),
            # pode diferir de pld_entered_at (quando entrou na fila PLD no Athena).
            "competencia_at": f"{ano}-{mes_num:02d}-01T00:00:00",
            "_source": "planilha",
            "_razao_social": str(row[4]) if row[4] else None,
        })

    return items


def mesclar(pep_history: dict, novos: list[dict]) -> tuple[dict, int, int]:
    """Adiciona casos novos sem duplicar. Retorna (history, adicionados, atualizados)."""
    existentes = {it["draft_membership_id"]: i for i, it in enumerate(pep_history["items"])}
    adicionados = 0
    atualizados = 0

    for it in novos:
        did = it["draft_membership_id"]
        if did in existentes:
            idx = existentes[did]
            atual = pep_history["items"][idx]
            changed = False
            # Atualiza status:
            # - planilha tem prioridade sobre Athena para falso_positivo (reprocessamento)
            # - planilha atualiza casos em_andamento para status final
            should_update = (
                it["status_pepito"] == "falso_positivo"
                or (atual["status_pepito"] == "em_andamento" and it["status_pepito"] != "em_andamento")
            )
            if should_update:
                pep_history["items"][idx]["status_pepito"] = it["status_pepito"]
                pep_history["items"][idx]["raw_status"] = it["raw_status"]
                pep_history["items"][idx]["_source"] = "planilha"
                changed = True
            # Adiciona competencia_at (não sobrescreve pld_entered_at)
            if "competencia_at" not in atual and it.get("competencia_at"):
                pep_history["items"][idx]["competencia_at"] = it["competencia_at"]
                pep_history["items"][idx]["_source"] = "planilha"
                changed = True
            if changed:
                atualizados += 1
        else:
            pep_history["items"].append(it)
            existentes[did] = len(pep_history["items"]) - 1
            adicionados += 1

    # Recalcula meta
    c = Counter(it["status_pepito"] for it in pep_history["items"])
    pep_history["_meta"]["by_status"] = dict(c)
    pep_history["_meta"]["total"] = len(pep_history["items"])
    pep_history["_meta"]["planilha_merged_at"] = datetime.now(timezone.utc).isoformat()
    pep_history["_meta"]["planilha_added"] = adicionados
    pep_history["_meta"]["planilha_updated"] = atualizados

    return pep_history, adicionados, atualizados


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--mes", type=int, default=None)
    parser.add_argument("--ano", type=int, default=None)
    args = parser.parse_args()

    print("=== Baixando CONTROLE DE ANÁLISES.xlsx ===")
    xlsx_bytes = _baixar_xlsx()

    print(f"=== Lendo aba '{ABA_PEP_ONB}' ===")
    headers, rows = _ler_aba(xlsx_bytes)
    print(f"  {len(rows)} linhas, {len(headers)} colunas")

    filtro_info = ""
    if args.mes:
        filtro_info += f" mês={args.mes}"
    if args.ano:
        filtro_info += f" ano={args.ano}"

    print(f"=== Processando{filtro_info} ===")
    itens = processar(rows, filtro_mes=args.mes, filtro_ano=args.ano)
    print(f"  {len(itens)} itens válidos da planilha")

    c = Counter(it["status_pepito"] for it in itens)
    for status, n in sorted(c.items(), key=lambda x: -x[1]):
        print(f"    {status}: {n}")

    if args.dry_run:
        print("\n[DRY RUN] Nenhuma alteração gravada.")
        return

    print("\n=== Mesclando com pep-history.json ===")
    with open(PEP_HISTORY_PATH) as f:
        pep_history = json.load(f)

    print(f"  pep-history antes: {pep_history['_meta']['total']} items")
    pep_history, adicionados, atualizados = mesclar(pep_history, itens)

    with open(PEP_HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(pep_history, f, ensure_ascii=False, indent=2)

    print(f"  + {adicionados} adicionados, {atualizados} atualizados")
    print(f"  pep-history depois: {pep_history['_meta']['total']} items")
    print(f"  by_status: {pep_history['_meta']['by_status']}")
    print("✓ pep-history.json atualizado")


if __name__ == "__main__":
    main()
